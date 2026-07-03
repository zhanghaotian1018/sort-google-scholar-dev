#!/usr/bin/env python3

# -*- coding: utf-8 -*-
"""
This code creates a database with a list of publications data from Google
Scholar.
The data acquired from GS is Title, Citations, Links and Rank.
It is useful for finding relevant papers by sorting by the number of citations
This example will look for the top 100 papers related to the keyword,
so that you can rank them by the number of citations

As output this program will plot the number of citations in the Y axis and the
rank of the result in the X axis. It also, optionally, export the database to
a .csv file.


"""

import requests
import datetime
import argparse
import csv
from bs4 import BeautifulSoup
import matplotlib.pyplot as plt
import pandas as pd
from time import sleep
import random
import re
import logging
import sys
from pathlib import Path
from io import StringIO
<<<<<<< HEAD
import urllib.parse
=======
>>>>>>> 5dbb8ec (first commit)

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Removed Python 2 compatibility for raw_input; using input() directly

# Default Parameters
KEYWORD = "machine learning"  # Default argument if command line is empty
NRESULTS = 100  # Fetch 100 articles
CSVPATH = Path.cwd()  # Default path as current working directory
SAVECSV = True
SORTBY = "Citations"
PLOT_RESULTS = False
STARTYEAR = None
now = datetime.datetime.now()
ENDYEAR = now.year  # Current year
DEBUG = False  # debug mode
MAX_CSV_FNAME = 255
LANG = "All"


# Websession Parameters
GSCHOLAR_URL = "https://scholar.google.com/scholar?start={}&q={}&hl=en&as_sdt=0,5"
YEAR_RANGE = ""  # &as_ylo={start_year}&as_yhi={end_year}'
# GSCHOLAR_URL_YEAR = GSCHOLAR_URL+YEAR_RANGE
STARTYEAR_URL = "&as_ylo={}"
ENDYEAR_URL = "&as_yhi={}"
LANG_URL = "&lr={}"

ROBOT_KW = ["unusual traffic from your computer network", "not a robot"]

<<<<<<< HEAD
# Browser-like User-Agent to avoid 429 blocks
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/125.0.0.0 Safari/537.36"
)

# Chromedriver path (selenium-manager downloads here)
CHROMEDRIVER_PATH = str(Path.home() / ".cache" / "selenium" / "chromedriver" / "win64")

=======
>>>>>>> 5dbb8ec (first commit)
# Set up logging
logging.basicConfig(level=logging.INFO)

# Initialize module logger
logger = logging.getLogger(__name__)


def get_command_line_args():
    # Command line arguments
    parser = argparse.ArgumentParser(description="Arguments")
    parser.add_argument(
        "kw",
<<<<<<< HEAD
        nargs='*',  # Allow multiple positional arguments
        help="""Keyword to be searched. Use double quote followed by simple quote to search for an exact keyword. Example: "'exact keyword'" """,
        default=[KEYWORD],
=======
        type=str,
        help="""Keyword to be searched. Use double quote followed by simple quote to search for an exact keyword. Example: "'exact keyword'" """,
        default=KEYWORD,
>>>>>>> 5dbb8ec (first commit)
    )
    parser.add_argument(
        "--sortby",
        type=str,
        help='Column to be sorted by. Default is by the columns "Citations", i.e., it will be sorted by the number of citations. If you want to sort by citations per year, use --sortby "cit/year"',
    )
    parser.add_argument(
        "--langfilter",
        nargs="+",
        type=str,
        help="Only languages listed are permitted to pass the filter. List of supported language codes: zh-CN, zh-TW, nl, en, fr, de, it, ja, ko, pl, pt, es, tr",
    )

    parser.add_argument(
        "--nresults",
        type=int,
        help="Number of articles to search on Google Scholar. Default is 100. (carefull with robot checking if value is too high)",
    )
    parser.add_argument(
        "--csvpath",
        type=str,
        help="Path to save the exported csv file. By default it is the current folder",
    )
    parser.add_argument(
        "--notsavecsv",
        action="store_true",
        help="By default results are going to be exported to a csv file. Select this option to just print results but not store them",
    )
    parser.add_argument(
        "--plotresults",
        action="store_true",
        help="Use this flag in order to plot the results with the original rank in the x-axis and the number of citaions in the y-axis. Default is False",
    )
    parser.add_argument(
        "--startyear", type=int, help="Start year when searching. Default is None"
    )
    parser.add_argument(
        "--endyear", type=int, help="End year when searching. Default is current year"
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Debug mode. Used for unit testing. It will get pages stored on web archive",
    )
    parser.add_argument(
        "--xlsx-only",
        action="store_true",
        help="Save only an Excel (.xlsx) file and skip creating a CSV. If openpyxl is missing, will fall back to CSV.",
    )

    # Parse and read arguments and assign them to variables if exists
    args, _ = parser.parse_known_args()

    # Check if no arguments were provided and print help if so
    if len(sys.argv) == 1:
        parser.print_help()
        sys.exit(0)

    keyword = KEYWORD
    if args.kw:
<<<<<<< HEAD
        # Join all keyword parts back together with spaces
        keyword = " ".join(args.kw)
=======
        keyword = args.kw
>>>>>>> 5dbb8ec (first commit)

    nresults = NRESULTS
    if args.nresults:
        nresults = args.nresults

    csvpath = CSVPATH
    if args.csvpath:
        csvpath = args.csvpath

    save_csv = SAVECSV
    if args.notsavecsv:
        save_csv = False

    sortby = SORTBY
    if args.sortby:
        sortby = args.sortby

    langfilter = LANG
    if args.langfilter:
        langfilter = args.langfilter

    plot_results = False
    if args.plotresults:
        plot_results = True

    start_year = STARTYEAR
    if args.startyear:
        start_year = args.startyear

    end_year = ENDYEAR
    if args.endyear:
        end_year = args.endyear

    debug = DEBUG
    if args.debug:
        debug = True

    xlsx_only = False
    if getattr(args, "xlsx_only", False):
        xlsx_only = True

    return (
        keyword,
        nresults,
        save_csv,
        csvpath,
        sortby,
        langfilter,
        plot_results,
        start_year,
        end_year,
        xlsx_only,
        debug,
    )


def get_citations(content: str) -> int:
    """Extract number of citations from content using regex."""
    match = re.search(r"Cited by (\d+)", content)
    return int(match.group(1)) if match else 0


def get_year(content: str) -> int:
    """Extract publication year from content using regex."""
    match = re.search(r"\b(19|20)\d{2}\b", content)
    return int(match.group(0)) if match else 0


def setup_driver() -> webdriver.Chrome:
    logger.info("Initializing WebDriver")
    chrome_options = Options()
    chrome_options.add_argument("disable-infobars")
<<<<<<< HEAD
    # Explicitly set service to let selenium-manager find chromedriver
    try:
        from selenium.webdriver.chrome.service import Service
        service = Service(CHROMEDRIVER_PATH)
        driver = webdriver.Chrome(service=service, options=chrome_options)
    except Exception:
        driver = webdriver.Chrome(options=chrome_options)
=======
    driver = webdriver.Chrome(options=chrome_options)
>>>>>>> 5dbb8ec (first commit)
    return driver


def clean_text(text: str) -> str:
    """Clean and normalize extracted text by removing special characters and extra whitespace."""
    if not text:
        return ""
    # Replace non-breaking spaces and other special whitespace characters
    text = text.replace("\xa0", " ").replace("\u200b", "").replace("\r", " ").replace("\n", " ")
    # Remove extra whitespace
    text = " ".join(text.split())
    return text.strip()


def get_author(content: str) -> str:
    """Extract the author string from content."""
    clean_content = clean_text(content)
    return clean_content.split(" - ")[0] if clean_content else ""


def get_venue_and_publisher(metadata_text: str) -> tuple:
    """
    Extract venue and publisher from metadata text.
    Google Scholar metadata format: Author - Year - Publisher/Journal/Venue
    
    Since Google Scholar doesn't explicitly separate Publisher and Venue,
    we treat the third part as the main publication source (Publisher),
    and extract Venue information if it contains journal/conference keywords.
    """
    if not metadata_text:
        return "", ""
    
    parts = metadata_text.split("-")
    
    venue = ""
    publisher = ""
    
    # Format: Author - Year - Publisher
    # [0] = Author, [1] = Year, [2+] = Publisher (might contain dashes)
    if len(parts) >= 3:
        # Year is in parts[1]
        # Everything after that is publisher/journal/venue
        publication_source = "-".join(parts[2:]).strip()
        publication_source = clean_text(publication_source)
        
        # Check if it's a venue/journal (contains typical keywords)
        venue_keywords = [
            "journal",
            "review",
            "proceedings",
            "conference",
            "symposium",
            "workshop",
            "transaction",
        ]
        
        publication_lower = publication_source.lower()
        is_venue = any(keyword in publication_lower for keyword in venue_keywords)
        
        if is_venue:
            # If it looks like a venue/journal, put it in both columns
            venue = publication_source
            publisher = publication_source
        else:
            # If it's just a publisher (like "Springer", "books.google.com")
            publisher = publication_source
            venue = ""  # Leave venue empty for non-journal publications
    
    return venue, publisher


def extract_venue_from_publisher(publisher: str) -> str:
    """
    Try to extract venue information from publisher text.
    This is kept for backward compatibility.
    """
    return ""  # Already handled in get_venue_and_publisher


def get_element(driver, xpath: str, attempts: int = 5, _count: int = 0):
    """Safely find an element by xpath with retries using updated selenium API."""
    try:
        return driver.find_element(By.XPATH, xpath)
    except Exception:
        if _count < attempts:
            sleep(random.uniform(0.5, 3))
            return get_element(driver, xpath, attempts=attempts, _count=_count + 1)
        logger.error("Element not found after %s attempts: %s", attempts, xpath)
        return None


def get_content_with_selenium(url):
    if "driver" not in globals():
        global driver
        driver = setup_driver()
    driver.get(url)

    while True:
        # Wait for a specific element that indicates the page has loaded
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )

        # Get the body element
        el = driver.find_element(By.TAG_NAME, "body")

        c = el.get_attribute("innerHTML")
        if any(kw in el.text for kw in ROBOT_KW):
<<<<<<< HEAD
            logger.warning("CAPTCHA detected. Please solve it in the browser window, then press Enter to continue.")
            try:
                input("Press Enter when CAPTCHA is solved...")
            except EOFError:
                logger.error("CAPTCHA requires manual interaction but stdin is not available (e.g., piped mode). Exiting.")
                sys.exit(1)
=======
            input("Solve captcha manually and press enter here to continue...")
>>>>>>> 5dbb8ec (first commit)
        else:
            break

    return c.encode("utf-8")


def format_strings(strings):
    if len(strings) == 1:
        return f"lang_{strings[0]}"
    else:
        return "%7C".join(f"lang_{s}" for s in strings)


def get_pdf_link(div):
    """Extract PDF link from the Google Scholar result if available"""
    try:
        pdf_div = div.find("div", {"class": "gs_ggs gs_fl"})
        if pdf_div:
            a_tag = pdf_div.find("a")
            if a_tag:
                return a_tag.get("href")
    except:
        pass
    return None


def main():
    # Get command line arguments
    (
        keyword,
        number_of_results,
        save_database,
        path,
        sortby_column,
        langfilter,
        plot_results,
        start_year,
        end_year,
        xlsx_only,
        debug,
    ) = get_command_line_args()

    logger.info(
        f"Running with parameters: Keyword: {keyword}, Number of results: {number_of_results}, Save database: {save_database}, Path: {path}, Sort by: {sortby_column}, Permitted Languages: {langfilter}, Plot results: {plot_results}, Start year: {start_year}, End year: {end_year}, Debug: {debug}"
    )

    # Create main URL based on command line arguments
    if start_year:
        GSCHOLAR_MAIN_URL = GSCHOLAR_URL + STARTYEAR_URL.format(start_year)
    else:
        GSCHOLAR_MAIN_URL = GSCHOLAR_URL

    if end_year != now.year:
        GSCHOLAR_MAIN_URL = GSCHOLAR_MAIN_URL + ENDYEAR_URL.format(end_year)

    if langfilter != "All":
        formatted_filters = format_strings(langfilter)
        GSCHOLAR_MAIN_URL = GSCHOLAR_MAIN_URL + LANG_URL.format(formatted_filters)

    if debug:
        GSCHOLAR_MAIN_URL = "https://web.archive.org/web/20210314203256/" + GSCHOLAR_URL

    # Start new session
    session = requests.Session()
<<<<<<< HEAD
    session.headers.update({
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
    })
=======
    # headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
>>>>>>> 5dbb8ec (first commit)

    # Variables
    links = []
    title = []
    citations = []
    year = []
    author = []
    venue = []
    publisher = []
    content = []  # Add new list for content
    pdf_links = []  # New list for PDF links
    rank = [0]

    # Get content from number_of_results URLs
    for n in range(0, number_of_results, 10):
        # if start_year is None:
<<<<<<< HEAD
        # Properly encode the keyword for URL
        encoded_keyword = urllib.parse.quote(keyword)
        url = GSCHOLAR_MAIN_URL.format(str(n), encoded_keyword)
=======
        url = GSCHOLAR_MAIN_URL.format(str(n), keyword.replace(" ", "+"))
>>>>>>> 5dbb8ec (first commit)
        if debug:
            logger.debug("Opening URL: %s", url)
        # else:
        #    url=GSCHOLAR_URL_YEAR.format(str(n), keyword.replace(' ','+'), start_year=start_year, end_year=end_year)

        logger.info("Loading next %d results", n + 10)
        page = session.get(url)  # , headers=headers)
        page.encoding = 'utf-8'  # Ensure proper encoding from the start
<<<<<<< HEAD
        c = page.content.decode("utf-8", errors="ignore")
        if any(kw in c for kw in ROBOT_KW):
=======
        c = page.content
        if any(kw in c.decode("utf-8", errors="ignore") for kw in ROBOT_KW):
>>>>>>> 5dbb8ec (first commit)
            logger.warning("Robot check detected, using Selenium fallback")
            try:
                c = get_content_with_selenium(url)
            except Exception as e:
                logger.exception(
                    "Failed to fetch content with Selenium for URL: %s", url
                )

        # Create parser with explicit encoding
        soup = BeautifulSoup(c, "html.parser")

        # Get stuff
<<<<<<< HEAD
        mydivs = soup.find_all("div", {"class": "gs_or"})
=======
        mydivs = soup.findAll("div", {"class": "gs_or"})
>>>>>>> 5dbb8ec (first commit)
        for div in mydivs:
            try:
                links.append(div.find("h3").find("a").get("href"))
            except:  # catch *all* exceptions
                links.append("Look manually at: " + url)

            try:
                title.append(clean_text(div.find("h3").find("a").text))
            except:
                title.append("Could not catch title")

            try:
                citations.append(get_citations(str(div.format_string)))
            except:
                logger.warning(
                    "Number of citations not found for %s. Appending 0", title[-1]
                )
                citations.append(0)

            try:
                year.append(get_year(div.find("div", {"class": "gs_a"}).text))
            except:
                logger.warning("Year not found for %s, appending 0", title[-1])
                year.append(0)

            try:
                author.append(clean_text(get_author(div.find("div", {"class": "gs_a"}).text)))
            except:
                author.append("Author not found")

            try:
                metadata = div.find("div", {"class": "gs_a"}).text
                venue_val, publisher_val = get_venue_and_publisher(metadata)
                venue.append(venue_val)
                publisher.append(publisher_val)
            except:
                venue.append("Venue not found")
                publisher.append("Publisher not found")

            try:
                content_div = div.find("div", {"class": "gs_rs"})
                if content_div:
                    # Get the text content and clean it
                    content_text = content_div.text
                    content_text = clean_text(content_text)
                    # Remove the ellipsis marker that Google Scholar adds
                    content_text = content_text.replace("… ", "").replace(" …", "").replace("…", "")
                    content.append(content_text if content_text else "Content not found")
                else:
                    content.append("Content not found")
            except:
                content.append("Content not found")

            # Extract PDF link
            pdf_links.append(clean_text(get_pdf_link(div) or "No PDF link"))

            rank.append(rank[-1] + 1)

        # Delay
        sleep(random.uniform(0.5, 3))

    # Create a dataset and sort by the number of citations
    data = pd.DataFrame(
        list(
            zip(
                author,
                title,
                citations,
                year,
                publisher,
                venue,
                content,
                links,
                pdf_links,
            )
        ),
        index=rank[1:],
        columns=[
            "Author",
            "Title",
            "Citations",
            "Year",
            "Publisher",
            "Venue",
            "Content",
            "Source",
            "PDF",
        ],
    )
    data.index.name = "Rank"

    # Avoid years that are higher than the current year by clipping it to end_year
<<<<<<< HEAD
    divisor = end_year + 1 - data["Year"].clip(upper=end_year)
    divisor = divisor.replace(0, 1)  # avoid division by zero
    data["cit/year"] = (data["Citations"] / divisor).round(0).astype(int)
=======
    data["cit/year"] = data["Citations"] / (
        end_year + 1 - data["Year"].clip(upper=end_year)
    )
    data["cit/year"] = data["cit/year"].round(0).astype(int)
>>>>>>> 5dbb8ec (first commit)

    # Sort by the selected columns, if exists
    try:
        data_ranked = data.sort_values(by=sortby_column, ascending=False)
    except Exception as e:
        logger.warning(
            "Sort column '%s' not found. Falling back to 'Citations'", sortby_column
        )
        data_ranked = data.sort_values(by="Citations", ascending=False)
        logger.debug("Sorting error details: %s", e)

    # Print data
    logger.info("Results:\n%s", data_ranked.to_string())

    # Plot by citation number
    if plot_results:
        plt.plot(rank[1:], citations, "*")
        plt.ylabel("Number of Citations")
        plt.xlabel("Rank of the keyword on Google Scholar")
        plt.title("Keyword: " + keyword)
        plt.show()

    # Save results
    if save_database:
        csv_file_name = f"{keyword.replace(' ', '_').replace(':', '_')}.csv"
        csv_path = Path(path) / csv_file_name
        # Truncate filename if too long
        if len(csv_path.name) > MAX_CSV_FNAME:
            csv_path = csv_path.with_name(csv_path.name[:MAX_CSV_FNAME])

        # If user requested xlsx-only, try to produce only the Excel file
        if xlsx_only:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment

                xlsx_path = csv_path.with_suffix('.xlsx')
                with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
                    data_ranked.to_excel(writer, sheet_name='Results', index=True)
                    worksheet = writer.sheets['Results']

                    column_widths = {
                        'A': 6,   # Rank
                        'B': 20,  # Author
                        'C': 35,  # Title
                        'D': 12,  # Citations
                        'E': 8,   # Year
                        'F': 25,  # Publisher
                        'G': 25,  # Venue
                        'H': 60,  # Content
                        'I': 40,  # Source
                        'J': 30,  # PDF
                        'K': 12,  # cit/year
                    }

                    for col, width in column_widths.items():
                        worksheet.column_dimensions[col].width = width

                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')

                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

                logger.info("Results saved to Excel: %s", xlsx_path)
            except ImportError:
                logger.warning("openpyxl not installed — falling back to CSV output")
                data_ranked.to_csv(csv_path, encoding="utf-8-sig", index=True, quoting=csv.QUOTE_ALL, lineterminator='\n')
                logger.info("Results saved to %s", csv_path)
        else:
            # Default behaviour: write CSV and also try to write Excel for better viewing
            data_ranked.to_csv(csv_path, encoding="utf-8-sig", index=True, quoting=csv.QUOTE_ALL, lineterminator='\n')
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment

                xlsx_path = csv_path.with_suffix('.xlsx')
                with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
                    data_ranked.to_excel(writer, sheet_name='Results', index=True)
                    worksheet = writer.sheets['Results']

                    column_widths = {
                        'A': 6,   # Rank
                        'B': 20,  # Author
                        'C': 35,  # Title
                        'D': 12,  # Citations
                        'E': 8,   # Year
                        'F': 25,  # Publisher
                        'G': 25,  # Venue
                        'H': 60,  # Content
                        'I': 40,  # Source
                        'J': 30,  # PDF
                        'K': 12,  # cit/year
                    }

                    for col, width in column_widths.items():
                        worksheet.column_dimensions[col].width = width

                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')

                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

                logger.info("Results saved to both CSV and Excel: %s, %s", csv_path, xlsx_path)
            except ImportError:
                logger.info("openpyxl not installed. Results saved to CSV only: %s", csv_path)

            logger.info("Results saved to %s", csv_path)


if __name__ == "__main__":
    main()
